from tkinter import *
import openpyxl  
import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email import encoders
from tkinter import messagebox
import mimetypes
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.text import MIMEText
import ctypes
from google_auth_oauthlib.flow import InstalledAppFlow
from tkinter import filedialog
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from email.encoders import encode_base64
import argparse
import base64
import os.path

#######################################################################################################################################################  

def excel_file_handler():
    Path=lab.cget("text")
    Accno = path1.get()
    Valid = path2.get()
    Name = path3.get()
    re_d=int(path4.get())
    f_name=path5.get()

    
    user_excel_file = openpyxl.load_workbook(filename = Path)                                                    
    sheet_names = user_excel_file.sheetnames   
    for current_sheet in sheet_names:         
        active_sheet = user_excel_file[current_sheet]         
        for row in active_sheet.rows:      
            for i in row:
                if i.value == Accno:         
                    Access_card_no = str(i)            
                    get_address_of_cell = Access_card_no.split('.')   
                    length_of_list = len(get_address_of_cell)
                    Address = get_address_of_cell[length_of_list - 1].split('>') 
                    Address_of_Access_card_no = Address[0]          
                    data_of_Access_card_no = active_sheet[Address_of_Access_card_no[0]]       
                    data_of_Access_card_no = list(data_of_Access_card_no)
                    
                elif i.value == Valid:
                    Valid_till = str(i)             
                    get_address_of_cell = Valid_till.split('.')   
                    length_of_list = len(get_address_of_cell)         
                    Address = get_address_of_cell[length_of_list - 1].split('>')   
                    Address_of_Valid_till = Address[0]           
                    data_of_Valid_till = active_sheet[Address_of_Valid_till[0]]     
                    data_of_Valid_till = list(data_of_Valid_till)           
                    print(data_of_Valid_till)

                elif i.value == Name:
                    Name_str = str(i)             
                    get_address_of_cell = Name_str.split('.')   
                    length_of_list = len(get_address_of_cell)         
                    Address = get_address_of_cell[length_of_list - 1].split('>')  
                    Address_of_Name_str = Address[0]           
                    data_of_Name_str = active_sheet[Address_of_Name_str[0]]      
                    data_of_Name_str = list(data_of_Name_str)       


    def change(list1,list2):
        change_list = [(list1[i],list2[i]) for i in range(0, len(list1))]
        return change_list
    valid_till_name=change(data_of_Valid_till,data_of_Name_str)
    

    dictionary = {}
    for key in data_of_Access_card_no:
        for value in valid_till_name:
            dictionary[key] = value
            valid_till_name.remove(value)
            break

        
    today = datetime.datetime.today()   
    today = today.replace(hour=0, minute=0, second=0, microsecond=0)
    final_dictionary = {}

    
    for key in dictionary:
        temp_value = dictionary[key][0].value      
        if type(temp_value) == type(today):      
            Valid_till_date = temp_value    
            Remaining_days = Valid_till_date - today   
            Remaining_days = Remaining_days.days 
            
            if ((Remaining_days >0)&(Remaining_days <= re_d)):                    
                final_dictionary[key] = (Remaining_days,dictionary[key][1])
                print(type(Remaining_days))

                workbook = openpyxl.Workbook()             
                sheet = workbook.active                    
                sheet["A1"] = "Access card no"            
                sheet["B1"] = "Name"
                sheet["C1"] = "Remaining days"
                                    
                i = 2                                      
                for key in final_dictionary:               
                    sheet["A" + str(i)] = key.value    
                    sheet["B" + str(i)] = final_dictionary[key][1].value
                    sheet["c" + str(i)] = final_dictionary[key][0]
                    i += 1                             
                workbook.save(filename=f_name)

    messagebox.showinfo('DONE', 'Succesfully Completed')

#######################################################################################################################################################  

def sendmail():

    f_name=path5.get()
    
    add=path6.get()
    tit=path7.get()
   
    mess=path8.get()
    
    def create_message(to, subject, message_text):
      """Create a message for an email.

      Args:
        sender: Email address of the sender.
        to: Email address of the receiver.
        subject: The subject of the email message.
        message_text: The text of the email message.

      Returns:
        An object containing a base64url encoded email object.
      """
      message = MIMEText(message_text)
      message['to'] = to
    #  message['from'] = sender
      message['subject'] = subject
      return {'raw': base64.urlsafe_b64encode(message.as_string().encode()).decode()}


    def create_message_with_attachment(to, subject, message_text, file):
      """Create a message for an email.

      Args:
        sender: Email address of the sender.
        to: Email address of the receiver.
        subject: The subject of the email message.
        message_text: The text of the email message.
        file: The path to the file to be attached.

      Returns:
        An object containing a base64url encoded email object.
      """
      message = MIMEMultipart()
      message['to'] = to
    #  message['from'] = sender
      message['subject'] = subject

      msg = MIMEText(message_text)
      message.attach(msg)

      content_type, encoding = mimetypes.guess_type(file)

      if content_type is None or encoding is not None:
        content_type = 'application/octet-stream'
      main_type, sub_type = content_type.split('/', 1)
      if main_type == 'text':
        fp = open(file, 'r')
        msg = MIMEText(fp.read(), _subtype=sub_type)
        fp.close()
      elif main_type == 'image':
        fp = open(file, 'rb')
        msg = MIMEImage(fp.read(), _subtype=sub_type)
        fp.close()
      elif main_type == 'audio':
        fp = open(file, 'rb')
        msg = MIMEAudio(fp.read(), _subtype=sub_type)
        fp.close()
      else:
        fp = open(file, 'rb')
        msg = MIMEBase(main_type, sub_type)
        msg.set_payload(fp.read())
        encode_base64(msg)
        fp.close()

      filename = f_name
      msg.add_header('Content-Disposition', 'attachment', filename=filename)
      message.attach(msg)

      return {'raw': base64.urlsafe_b64encode(message.as_string().encode()).decode()}

    def send_message(service, user_id, message):
      """Send an email message.

      Args:
        service: Authorized Gmail API service instance.
        user_id: User's email address. The special value "me"
        can be used to indicate the authenticated user.
        message: Message to be sent.

      Returns:
        Sent Message.
      """
      message = (service.users().messages().send(userId=user_id, body=message)
                   .execute())
      print(f"email sent")


    # If modifying these scopes, delete the file token.json.
    SCOPES = ['https://www.googleapis.com/auth/gmail.send']

    def make_googleapi_verification():
        """Shows basic usage of the Gmail API.
        Lists the user's Gmail labels.
        """
        creds = None
        # The file token.json stores the user's access and refresh tokens, and is
        # created automatically when the authorization flow completes for the first
        # time.
        if os.path.exists('token.json'):
            creds = Credentials.from_authorized_user_file('token.json', SCOPES)
        # If there are no (valid) credentials available, let the user log in.
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    'credentials.json', SCOPES)
                creds = flow.run_local_server(port=0)
            # Save the credentials for the next run
            with open('token.json', 'w') as token:
                token.write(creds.to_json())

        return build('gmail', 'v1', credentials=creds)


    def send_email(address, title, message, attachment=''):
        service = make_googleapi_verification()
        if attachment == '':
            message = create_message(address, title, message)
        else:
            message = create_message_with_attachment(address, title, message, attachment)

        send_message(service, "me", message)


    def email_parse_argument():
        parser = argparse.ArgumentParser()
        parser.add_argument("-a", "--address", required=True, help="Receiver email address")
        parser.add_argument("-t", "--title", required=True, help="Title of the email")
        parser.add_argument("-m", "--message", required=True, help="Email message")
        parser.add_argument("-f", "--attach", default='', help="Attachment file name")
        return parser.parse_args()

    def email_main():
        args = email_parse_argument()
        send_email(args.address, args.title, args.message, args.attach)

##    if __name__ == '__main__':
##        email_main()

    send_email(add, tit, mess, f_name)
               
#######################################################################################################################################################      

root = Tk()
root.title('Expired Remider')
#root.geometry("900x650")
root.geometry("1000x650+%d+%d" %( ( (root.winfo_screenwidth() / 2.) - (900 / 2.) ), ( (root.winfo_screenheight() / 2.) - (700 / 2.) ) ) )
root.config(background = "#5AB4E8")
path1 = StringVar()
path2 = StringVar()
path3 = StringVar()
path4=StringVar()
path5=StringVar()
path6=StringVar()
path7=StringVar()
path8=StringVar()

def browseFiles():
    filename = filedialog.askopenfilename(title = "Select a File",filetypes = (("Text files","*.xlsx*"),("all files","*.*")))
    lab.configure(text=filename)
    ctypes.windll.user32.MessageBoxW(0, "File Selected", filename, 0)
#######################################################################################################################################################    
lab = Label(root,text="Select Your Excel File",anchor="e",font=("Times New Roman",20),height=0,bg='#F8E828')
lab.grid(row=0,column=2 )

button_explore = Button(root,text = "Browse File",font=("Elephant",20),command =browseFiles)
button_explore.grid(row=1,column=2,sticky="ew")

#######################################################################################################################################################  

lab1 =Label(root,text="Enter the title of the Access card number  field : ",anchor="w",font=("Times New Roman",15),pady="10")
lab1.grid(row="3",column="1",sticky="ew")

txtName1 = Entry(root,font=("Times New Roman",15),textvariable = path1)
txtName1.grid(row="3",column="2",sticky="ew")

#######################################################################################################################################################  

lab2 =Label(root,text="Give the field name of Date: ",anchor="w",font=("Times New Roman",15),pady="10")
lab2.grid(row="4",column="1",sticky="ew")

txtName2 = Entry(root,font=("Times New Roman",15),textvariable = path2)
txtName2.grid(row="4",column="2",sticky="ew")

#######################################################################################################################################################  

lab3 =Label(root,text="Enter the field name of Customer_Name: ",anchor="w",font=("Times New Roman",15),pady="10")
lab3.grid(row="5",column="1",sticky="ew")

txtName3 = Entry(root,font=("Times New Roman",15),textvariable = path3)
txtName3.grid(row="5",column="2",sticky="ew")

#######################################################################################################################################################  

lab4 =Label(root,text="How many days before you get ALERT : ",anchor="w",font=("Times New Roman",15),pady="10")
lab4.grid(row="6",column="1",sticky="ew")

txtName4 = Entry(root,font=("Times New Roman",15),textvariable = path4)
txtName4.grid(row="6",column="2",sticky="ew")

#######################################################################################################################################################  

lab4 =Label(root,text="Choose one name for save your OUTPUT File : ",anchor="w",font=("Times New Roman",15),pady="10")
lab4.grid(row="7",column="1",sticky="ew")

txtName4 = Entry(root,font=("Times New Roman",15),textvariable = path5)
txtName4.grid(row="7",column="2",sticky="ew")

#######################################################################################################################################################  

But = Button(root,text ="Run",fg="black", font=("Elephant",20),bg='#FB00FF' , command = excel_file_handler)
But.grid(row="20",column="2",sticky="ew")

#######################################################################################################################################################  


lab5 =Label(root,text="Type SENDER E-Mail Address : ",anchor="w",font=("Times New Roman",15),pady="10")
lab5.grid(row="24",column="1",sticky="ew")

txtName5 = Entry(root,font=("Times New Roman",15),textvariable = path6)
txtName5.grid(row="24",column="2",sticky="ew")

#######################################################################################################################################################  

lab6 =Label(root,text="Type Your E-Mail Title : ",anchor="w",font=("Times New Roman",15),pady="10")
lab6.grid(row="25",column="1",sticky="ew")

txtName6 = Entry(root,font=("Times New Roman",15),show='*',textvariable = path7)
txtName6.grid(row="25",column="2",sticky="ew")

#######################################################################################################################################################  

lab7 =Label(root,text="Type Your E-Mail Message :  ",anchor="w",font=("Times New Roman",15),pady="10")
lab7.grid(row="26",column="1",sticky="ew")

txtName7 = Entry(root,font=("Times New Roman",15),textvariable = path8)
txtName7.grid(row="26",column="2",sticky="ew")

#######################################################################################################################################################  

But_mail = Button(root,text ="Send Mail",fg="black", font=("Elephant",20),bg='#938F93' , command = sendmail)
But_mail.grid(row="27",column="2",sticky="ew")

#######################################################################################################################################################  

button_exit = Button(root,text = "Exit",font=("Elephant",20),command = exit,bg='#FF2300')
button_exit.grid(row="21",column="2",sticky="ew")

root.mainloop()






    


