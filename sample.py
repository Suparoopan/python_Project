from tkinter import *
import openpyxl  
import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email import encoders
import mimetypes
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.text import MIMEText
import ctypes
from tkinter import filedialog

def excel_file_handler():
    Path=lab.cget("text")
    Accno = path1.get()
    Valid = path2.get()
    Name = path3.get()
    re_d=int(path4.get())
    f_name=path5.get()

    
    user_excel_file = openpyxl.load_workbook(filename = Path)                                                    
    sheet_names = user_excel_file.sheetnames   
    for current_sheet in sheet_names:         
        active_sheet = user_excel_file[current_sheet]         
        for row in active_sheet.rows:      
            for i in row:
                if i.value == Accno:         
                    Access_card_no = str(i)            
                    get_address_of_cell = Access_card_no.split('.')   
                    length_of_list = len(get_address_of_cell)
                    Address = get_address_of_cell[length_of_list - 1].split('>') 
                    Address_of_Access_card_no = Address[0]          
                    data_of_Access_card_no = active_sheet[Address_of_Access_card_no[0]]       
                    data_of_Access_card_no = list(data_of_Access_card_no)
                    
                elif i.value == Valid:
                    Valid_till = str(i)             
                    get_address_of_cell = Valid_till.split('.')   
                    length_of_list = len(get_address_of_cell)         
                    Address = get_address_of_cell[length_of_list - 1].split('>')   
                    Address_of_Valid_till = Address[0]           
                    data_of_Valid_till = active_sheet[Address_of_Valid_till[0]]     
                    data_of_Valid_till = list(data_of_Valid_till)           
                    print(data_of_Valid_till)

                elif i.value == Name:
                    Name_str = str(i)             
                    get_address_of_cell = Name_str.split('.')   
                    length_of_list = len(get_address_of_cell)         
                    Address = get_address_of_cell[length_of_list - 1].split('>')  
                    Address_of_Name_str = Address[0]           
                    data_of_Name_str = active_sheet[Address_of_Name_str[0]]      
                    data_of_Name_str = list(data_of_Name_str)       


    def change(list1,list2):
        change_list = [(list1[i],list2[i]) for i in range(0, len(list1))]
        return change_list
    valid_till_name=change(data_of_Valid_till,data_of_Name_str)
    

    dictionary = {}
    for key in data_of_Access_card_no:
        for value in valid_till_name:
            dictionary[key] = value
            valid_till_name.remove(value)
            break

        
    today = datetime.datetime.today()   
    today = today.replace(hour=0, minute=0, second=0, microsecond=0)
    final_dictionary = {}

    
    for key in dictionary:
        temp_value = dictionary[key][0].value      
        if type(temp_value) == type(today):      
            Valid_till_date = temp_value    
            Remaining_days = Valid_till_date - today   
            Remaining_days = Remaining_days.days 
            
            if ((Remaining_days >0)&(Remaining_days <= re_d)):                    
                final_dictionary[key] = (Remaining_days,dictionary[key][1])
                print(type(Remaining_days))

                workbook = openpyxl.Workbook()             
                sheet = workbook.active                    
                sheet["A1"] = "Access card no"            
                sheet["B1"] = "Name"
                sheet["C1"] = "Remaining days"
                                    
                i = 2                                      
                for key in final_dictionary:               
                    sheet["A" + str(i)] = key.value    
                    sheet["B" + str(i)] = final_dictionary[key][1].value
                    sheet["c" + str(i)] = final_dictionary[key][0]
                    i += 1                             
                workbook.save(filename=f_name)

    ctypes.windll.user32.MessageBoxW(0, "Succesfully Completed", "Done", 0)

def sendmail():
    f_name=path5.get()
    
    emailto=path8.get()
    username=''
    password=''
    
   

    if((username=='')&(password=='')):
        username=path6.get()
        password=path7.get()

        
        
        emailfrom = username
        
        fileToSend1 = [f_name]
        
        msg = MIMEMultipart()
        msg["From"] = emailfrom
        msg["To"] = emailto
        msg["Subject"] = "Hi, This gmail send by python. Here attach Three files."
        msg.preamble = "Hi, This gmail send by python. Here attach Three files."
        
        for fileToSend in fileToSend1:
            ctype, encoding = mimetypes.guess_type(fileToSend)
            if ctype is None or encoding is not None:
               ctype = "application/octet-stream"
        
            maintype, subtype = ctype.split("/", 1)
        
            if maintype == "text":
                fp = open(fileToSend)
                
                attachment = MIMEText(fp.read(), _subtype=subtype)
                fp.close()
            elif maintype == "image":
                fp = open(fileToSend, "rb")
                attachment = MIMEImage(fp.read(), _subtype=subtype)
                fp.close()
            elif maintype == "audio":
                fp = open(fileToSend, "rb")
                attachment = MIMEAudio(fp.read(), _subtype=subtype)
                fp.close()
            else:
                fp = open(fileToSend, "rb")
                attachment = MIMEBase(maintype, subtype)
                attachment.set_payload(fp.read())
                fp.close()
                encoders.encode_base64(attachment)
            attachment.add_header("Content-Disposition", "attachment", filename=fileToSend)
            msg.attach(attachment)
        
        server = smtplib.SMTP("smtp.gmail.com:587")
        server.starttls()
        server.login(username,password)
        server.sendmail(emailfrom, emailto, msg.as_string())
        server.quit()

        ctypes.windll.user32.MessageBoxW(0, "Enter the E-Mail Details", "Warning", 0)
        
        print('mail sent')

        return username,password;

    #else:
        
    

                
                                               
                
    

root = Tk()
root.title('Expired Remider')
#root.geometry("900x650")
root.geometry("1000x650+%d+%d" %( ( (root.winfo_screenwidth() / 2.) - (900 / 2.) ), ( (root.winfo_screenheight() / 2.) - (700 / 2.) ) ) )
root.config(background = "#5AB4E8")
path1 = StringVar()
path2 = StringVar()
path3 = StringVar()
path4=StringVar()
path5=StringVar()
path6=StringVar()
path7=StringVar()
path8=StringVar()

def browseFiles():
    filename = filedialog.askopenfilename(title = "Select a File",filetypes = (("Text files","*.xlsx*"),("all files","*.*")))
    lab.configure(text=filename)
    ctypes.windll.user32.MessageBoxW(0, "File Selected", filename, 0)
    
lab = Label(root,text="Select Your Excel File",font=("Times New Roman",20),height=0,bg='#F8E828')
lab.grid(row=0,column=2 )


button_explore = Button(root,text = "Browse File",font=("Elephant",20),command =browseFiles)
button_explore.grid(row=1,column=2,sticky="ew")


lab1 =Label(root,text="Enter the title of the Access card number  field : ",font=("Times New Roman",15),pady="10")
lab1.grid(row="3",column="1",sticky="ew")

txtName1 = Entry(root,font=("Times New Roman",15),textvariable = path1)
txtName1.grid(row="3",column="2",sticky="ew")

lab2 =Label(root,text="Give the field name of Date: ",font=("Times New Roman",15),pady="10")
lab2.grid(row="4",column="1",sticky="ew")

txtName2 = Entry(root,font=("Times New Roman",15),textvariable = path2)
txtName2.grid(row="4",column="2",sticky="ew")

lab3 =Label(root,text="Enter the field name of Customer_Name: ",font=("Times New Roman",15),pady="10")
lab3.grid(row="5",column="1",sticky="ew")

txtName3 = Entry(root,font=("Times New Roman",15),textvariable = path3)
txtName3.grid(row="5",column="2",sticky="ew")

lab4 =Label(root,text="How many days before you get ALERT : ",font=("Times New Roman",15),pady="10")
lab4.grid(row="6",column="1",sticky="ew")

txtName4 = Entry(root,font=("Times New Roman",15),textvariable = path4)
txtName4.grid(row="6",column="2",sticky="ew")

lab4 =Label(root,text="Choose one name for save your OUTPUT File : ",font=("Times New Roman",15),pady="10")
lab4.grid(row="7",column="1",sticky="ew")

txtName4 = Entry(root,font=("Times New Roman",15),textvariable = path5)
txtName4.grid(row="7",column="2",sticky="ew")

But = Button(root,text ="Run",fg="black", font=("Elephant",20),bg='#FB00FF' , command = excel_file_handler)
But.grid(row="20",column="2",sticky="ew")




lab5 =Label(root,text="Type your E-Mail USERNAME : ",font=("Times New Roman",15),pady="10")
lab5.grid(row="24",column="1",sticky="ew")

txtName5 = Entry(root,font=("Times New Roman",15),textvariable = path6)
txtName5.grid(row="24",column="2",sticky="ew")

lab6 =Label(root,text="Type Your E-Mail PASSWORD : ",font=("Times New Roman",15),pady="10")
lab6.grid(row="25",column="1",sticky="ew")

txtName6 = Entry(root,font=("Times New Roman",15),show='*',textvariable = path7)
txtName6.grid(row="25",column="2",sticky="ew")

lab7 =Label(root,text="Enter the sender E-Mail ADDRESS : ",font=("Times New Roman",15),pady="10")
lab7.grid(row="26",column="1",sticky="ew")

txtName7 = Entry(root,font=("Times New Roman",15),textvariable = path8)
txtName7.grid(row="26",column="2",sticky="ew")

But_mail = Button(root,text ="Send Mail",fg="black", font=("Elephant",20),bg='#938F93' , command = sendmail)
But_mail.grid(row="27",column="2",sticky="ew")


button_exit = Button(root,text = "Exit",font=("Elephant",20),command = exit,bg='#FF2300')
button_exit.grid(row="21",column="3",sticky="ew")

root.mainloop()






    


